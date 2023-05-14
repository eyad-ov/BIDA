import sys
import openpyxl
from Bio import SeqIO
from openpyxl.chart import BarChart, Reference

genetic_code = {"GCT": "A", "GCC": "A", "GCA": "A", "GCG": "A", "TGT": "C", "TGC": "C",
"GAT": "D", "GAC": "D",
"GAA": "E", "GAG": "E",
"TTT": "F", "TTC": "F",
"GGT": "G", "GGC": "G", "GGA": "G", "GGG": "G",
"CAT": "H", "CAC": "H",
"ATT": "I", "ATC": "I", "ATA": "I",
"AAA": "K", "AAG": "K",
"TTA": "L", "TTG": "L", "CTT": "L", "CTC": "L", "CTA": "L", "CTG": "L",
"ATG": "M",
"AAT": "N", "AAC": "N",
"CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P",
"CAA": "Q", "CAG": "Q",
"CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "AGA": "R", "AGG": "R",
"TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "AGT": "S", "AGC": "S",
"ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T",
"GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V",
"TGG": "W",
"TAT": "Y", "TAC": "Y",
"TAA": "*", "TAG": "*", "TGA": "*"}


wb = openpyxl.Workbook()
sheet = wb.active

z = sheet.cell(1,1)
z.value = "Triplet"
z = sheet.cell(1,2)
z.value = "Amino Acid"

row=2
col=1

for c in genetic_code:
 z = sheet.cell(row, col)
 z.value = c
 z = sheet.cell(row,col+1)
 z.value = genetic_code[c]
 row += 1

t_count = {}
aa_count = {}
fn = len(sys.argv) - 2
i = 2
while i <= fn:
    f = sys.argv[i]
    name = sys.argv[i+1]
    i += 2
    col += 2 
    z = sheet.cell(1,col)
    z.value = "Total " + name
    z = sheet.cell(1,col+1)
    z.value = "Percent "+ name
    row = 2
    handle = open(f, "r")
    for record in SeqIO.parse(handle, "fasta"):
        block = record.seq
        for j in range (0, len(block)-2):
         codon = block[j:j+3]
         if codon in genetic_code:
          protein = genetic_code[ codon ]
          if codon in t_count:
              t_count[str(codon)] += 1
          else:
              t_count[str(codon)] = 1
          if protein in aa_count:
              aa_count[protein] += 1
          else:
              aa_count[protein] = 1
    handle.close()
    for c in genetic_code:
      z = sheet.cell(row, col)
      total = t_count[c]
      z.value = total
      z = sheet.cell(row,col+1)
      percent = total / aa_count[genetic_code[c]]  * 100
      z.value = round(percent,2)
      row += 1
    
    aa_count = {}
    t_count = {}    

sheet2 = wb.create_sheet("Single Bar Charts")
i = 2
col = 4
n = 1
while i <= fn:
 chart = BarChart()
 chart.type = "col"
 chart.width = 35
 chart.height = 7.5
 chart.y_axis.scaling.max = 100
 chart.title = "Codon Usage " + sys.argv[i+1] 
 chart.y_axis.title = "Percentage Usage per Amino Acid"
 chart.x_axis.title = "Triplet"
 data = Reference(sheet, min_col=col, min_row=1, max_row=65)
 cats = Reference(sheet, min_col=1, min_row=2, max_row=65)
 chart.add_data(data, titles_from_data=True)
 chart.set_categories(cats)
 sheet2.add_chart(chart, "A" + str(n))
 i += 2
 col +=2
 n+=16

sheet3 = wb.create_sheet("Combined Bar Charts")
chart = BarChart()
chart.type = "col"
chart.width = 48
chart.height = 12
chart.y_axis.scaling.max = 100 
chart.y_axis.title = "Percentage Usage per Amino Acid"
chart.x_axis.title = "Triplet"

col = 4
num = fn//2
for k in range(num):
 data = Reference(sheet, min_col=col, min_row=1,max_row=65)
 chart.add_data(data, titles_from_data=True)
 col += 2 

cats = Reference(sheet, min_col=1, min_row=2, max_row=65)
chart.set_categories(cats)
sheet3.add_chart(chart, "A1")

wb.save(sys.argv[1] + ".xlsx")
